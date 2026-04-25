from datetime import datetime, timezone
from typing import Optional, List
from io import BytesIO
from bson import ObjectId
from app.database import get_database
from app.assets.service import get_asset_financial_data


async def get_comparison_report(user_id: str, asset_ids: List[str] = None) -> List[dict]:
    """Compare multiple assets side by side."""
    db = get_database()

    query = {"user_id": user_id}
    if asset_ids:
        query["_id"] = {"$in": [ObjectId(aid) for aid in asset_ids]}

    assets = []
    async for asset in db.assets.find(query):
        financial = await get_asset_financial_data(asset)
        assets.append({
            "id": str(asset["_id"]),
            "name": asset["name"],
            "type": asset["type"],
            "status": asset["status"],
            "initial_investment": asset["initial_investment"],
            "current_value": financial["current_value"],
            "accumulated_depreciation": financial["accumulated_depreciation"],
            "total_expenses": financial["total_expenses"],
            "total_income": financial["total_income"],
            "net_result": financial["net_result"],
            "roi": financial["roi"],
        })

    # Sort by ROI descending
    assets.sort(key=lambda x: x["roi"], reverse=True)
    return assets


async def get_period_report(
    user_id: str,
    start_date: datetime,
    end_date: datetime,
) -> dict:
    """Get financial report for a specific period."""
    db = get_database()

    # Transactions in period
    pipeline = [
        {
            "$match": {
                "user_id": user_id,
                "date": {"$gte": start_date, "$lte": end_date},
            }
        },
        {
            "$group": {
                "_id": "$type",
                "total": {"$sum": "$amount"},
                "count": {"$sum": 1},
            }
        },
    ]

    totals = {}
    async for doc in db.transactions.aggregate(pipeline):
        totals[doc["_id"]] = {"total": round(doc["total"], 2), "count": doc["count"]}

    # By asset in period
    asset_pipeline = [
        {
            "$match": {
                "user_id": user_id,
                "date": {"$gte": start_date, "$lte": end_date},
            }
        },
        {
            "$group": {
                "_id": "$asset_id",
                "total_expenses": {
                    "$sum": {
                        "$cond": [
                            {"$in": ["$type", ["mantenimiento", "operativo", "mejora"]]},
                            "$amount",
                            0,
                        ]
                    }
                },
                "total_income": {
                    "$sum": {
                        "$cond": [{"$eq": ["$type", "ingreso"]}, "$amount", 0]
                    }
                },
            }
        },
    ]

    by_asset = []
    async for doc in db.transactions.aggregate(asset_pipeline):
        asset = await db.assets.find_one({"_id": ObjectId(doc["_id"])})
        by_asset.append({
            "asset_id": doc["_id"],
            "asset_name": asset["name"] if asset else "Eliminado",
            "total_expenses": round(doc["total_expenses"], 2),
            "total_income": round(doc["total_income"], 2),
            "net": round(doc["total_income"] - doc["total_expenses"], 2),
        })

    total_expenses = sum(
        v["total"] for k, v in totals.items() if k in ["mantenimiento", "operativo", "mejora"]
    )
    total_income = totals.get("ingreso", {}).get("total", 0)

    return {
        "period": {
            "start": start_date.isoformat(),
            "end": end_date.isoformat(),
        },
        "summary": {
            "total_expenses": round(total_expenses, 2),
            "total_income": round(total_income, 2),
            "net_result": round(total_income - total_expenses, 2),
        },
        "by_type": totals,
        "by_asset": by_asset,
    }


async def get_top_assets(
    user_id: str,
    metric: str = "roi",
    limit: int = 10,
) -> List[dict]:
    """Get top assets ranked by a specific metric."""
    db = get_database()

    assets = []
    async for asset in db.assets.find({"user_id": user_id}):
        financial = await get_asset_financial_data(asset)
        assets.append({
            "id": str(asset["_id"]),
            "name": asset["name"],
            "type": asset["type"],
            "status": asset["status"],
            "initial_investment": asset["initial_investment"],
            "current_value": financial["current_value"],
            "total_expenses": financial["total_expenses"],
            "total_income": financial["total_income"],
            "roi": financial["roi"],
            "net_result": financial["net_result"],
        })

    # Sort by metric
    reverse = True
    if metric == "total_expenses":
        reverse = True  # Higher expenses first (most costly)
    elif metric == "roi":
        reverse = True

    assets.sort(key=lambda x: x.get(metric, 0), reverse=reverse)
    return assets[:limit]


# ─────────────────────────────── EXCEL EXPORT ───────────────────────────────


async def export_report_excel(user_id: str, report_type: str, **kwargs) -> BytesIO:
    """Generate Excel report."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active

    # Styles
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    if report_type == "comparison":
        ws.title = "Comparativa de Activos"
        headers = [
            "Activo", "Tipo", "Estado", "Inversión Inicial",
            "Valor Actual", "Depreciación", "Gastos Totales",
            "Ingresos Totales", "Resultado Neto", "ROI %"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        data = await get_comparison_report(user_id)
        for row_idx, asset in enumerate(data, 2):
            values = [
                asset["name"], asset["type"], asset["status"],
                asset["initial_investment"], asset["current_value"],
                asset["accumulated_depreciation"], asset["total_expenses"],
                asset["total_income"], asset["net_result"], asset["roi"],
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                if isinstance(val, float):
                    cell.number_format = '#,##0.00'

    elif report_type == "period":
        ws.title = "Reporte por Período"
        report = await get_period_report(user_id, kwargs["start_date"], kwargs["end_date"])

        # Summary section
        ws.cell(row=1, column=1, value="Reporte Financiero por Período").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Período: {report['period']['start']} - {report['period']['end']}")
        ws.cell(row=4, column=1, value="Total Gastos:").font = Font(bold=True)
        ws.cell(row=4, column=2, value=report["summary"]["total_expenses"]).number_format = '#,##0.00'
        ws.cell(row=5, column=1, value="Total Ingresos:").font = Font(bold=True)
        ws.cell(row=5, column=2, value=report["summary"]["total_income"]).number_format = '#,##0.00'
        ws.cell(row=6, column=1, value="Resultado Neto:").font = Font(bold=True)
        ws.cell(row=6, column=2, value=report["summary"]["net_result"]).number_format = '#,##0.00'

        # By asset detail
        headers = ["Activo", "Gastos", "Ingresos", "Neto"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row_idx, item in enumerate(report["by_asset"], 9):
            ws.cell(row=row_idx, column=1, value=item["asset_name"])
            ws.cell(row=row_idx, column=2, value=item["total_expenses"]).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=3, value=item["total_income"]).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=4, value=item["net"]).number_format = '#,##0.00'

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_length + 2, 12)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─────────────────────────────── PDF EXPORT ───────────────────────────────


async def export_report_pdf(user_id: str, report_type: str, **kwargs) -> BytesIO:
    """Generate PDF report."""
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter))
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.HexColor("#1a237e"),
    )

    if report_type == "comparison":
        elements.append(Paragraph("Comparativa de Activos — InfraControl", title_style))
        elements.append(Spacer(1, 0.3 * inch))

        data = await get_comparison_report(user_id)
        table_data = [
            ["Activo", "Tipo", "Inversión", "Valor Actual", "Gastos", "Ingresos", "ROI %"]
        ]
        for asset in data:
            table_data.append([
                asset["name"],
                asset["type"],
                f"${asset['initial_investment']:,.2f}",
                f"${asset['current_value']:,.2f}",
                f"${asset['total_expenses']:,.2f}",
                f"${asset['total_income']:,.2f}",
                f"{asset['roi']:.2f}%",
            ])

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f5f5f5")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table)

    elif report_type == "period":
        report = await get_period_report(user_id, kwargs["start_date"], kwargs["end_date"])
        elements.append(Paragraph("Reporte Financiero por Período — InfraControl", title_style))
        elements.append(Spacer(1, 0.2 * inch))
        elements.append(Paragraph(
            f"Período: {report['period']['start']} — {report['period']['end']}",
            styles["Normal"]
        ))
        elements.append(Spacer(1, 0.2 * inch))

        # Summary
        summary_data = [
            ["Métrica", "Monto"],
            ["Total Gastos", f"${report['summary']['total_expenses']:,.2f}"],
            ["Total Ingresos", f"${report['summary']['total_income']:,.2f}"],
            ["Resultado Neto", f"${report['summary']['net_result']:,.2f}"],
        ]
        summary_table = Table(summary_data, colWidths=[3 * inch, 3 * inch])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]))
        elements.append(summary_table)

    doc.build(elements)
    output.seek(0)
    return output
